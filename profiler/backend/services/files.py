"""Upload storage and document readers for CSV and Excel files."""

import io
import re
import warnings
from typing import Any, Dict, List, Optional
from uuid import uuid4

import pandas as pd
import polars as pl
from fastapi import HTTPException


FILE_STORE = {}
EXCEL_HEADER_SCAN_ROWS = 20
EXCEL_HEADER_SCAN_COLS = 1000


warnings.filterwarnings("ignore", message=r".*Could not determine dtype.*")


# In-memory file store used by follow-up sheet, filter, and matrix requests.
def store_file(contents: bytes, file_type: str, file_name: str) -> str:
    # Keep the raw upload in memory so later requests can re-read it for filters and matrices.
    file_id = str(uuid4())
    FILE_STORE[file_id] = {
        "contents": contents,
        "fileType": file_type,
        "fileName": file_name,
    }
    return file_id


def get_stored_file(file_id: str) -> Dict[str, Any]:
    stored = FILE_STORE.get(file_id)

    if stored is None:
        raise HTTPException(400, "File expired")

    # Older Excel-only entries were stored as raw bytes; keep this fallback during development.
    if isinstance(stored, bytes):
        return {
            "contents": stored,
            "fileType": "excel",
            "fileName": "uploaded-file",
        }

    return stored


def read_stored_dataframe(file_id: str, sheet_name: Optional[str] = None) -> pl.DataFrame:
    # CSVs are read as text first so mixed-type columns do not crash parsing.
    # Numeric/date profiling later uses safe casts where needed.
    stored = get_stored_file(file_id)
    contents = stored["contents"]

    if stored["fileType"] == "csv":
        return pl.read_csv(io.BytesIO(contents), infer_schema=False)

    return read_excel_dataframe(contents, sheet_name, stored["fileName"])


# Excel header detection handles workbooks whose true header row is not row 1.
def normalize_excel_header_value(value: Any) -> str:
    if value is None:
        return ""

    if isinstance(value, float) and value.is_integer():
        return str(int(value))

    return str(value).strip()


def make_unique_column_names(values: List[Any]) -> List[str]:
    seen = {}
    columns = []

    for index, value in enumerate(values, start=1):
        name = normalize_excel_header_value(value)

        if not name or re.match(r"^__UNNAMED__\d+$", name):
            name = f"Column {index}"

        count = seen.get(name, 0)
        seen[name] = count + 1

        if count:
            name = f"{name}_{count + 1}"

        columns.append(name)

    return columns


def is_numeric_header_value(value: str) -> bool:
    if not value:
        return False

    return re.fullmatch(r"[-+]?\d+(?:\.\d+)?(?:_\d+)?", value) is not None


def score_excel_header_row(values: List[Any]) -> float:
    normalized = [normalize_excel_header_value(value) for value in values]
    non_empty = [value for value in normalized if value]

    if not non_empty:
        return -1000

    text_like_count = sum(1 for value in non_empty if not is_numeric_header_value(value))
    numeric_count = len(non_empty) - text_like_count
    duplicate_count = len(non_empty) - len(set(non_empty))
    blank_count = len(normalized) - len(non_empty)

    return (
        len(non_empty) * 4
        + text_like_count * 3
        - numeric_count * 2
        - duplicate_count * 4
        - blank_count * 0.15
    )


def detect_excel_header_row(contents: bytes, sheet_name: Optional[str] = None) -> Dict[str, Any]:
    from openpyxl import load_workbook

    workbook = load_workbook(io.BytesIO(contents), read_only=True, data_only=True)
    worksheet = workbook[sheet_name] if sheet_name else workbook[workbook.sheetnames[0]]
    max_column = min(worksheet.max_column or 1, EXCEL_HEADER_SCAN_COLS)
    best_row_number = 1
    best_row = []
    best_score = -1000

    for row_number, row in enumerate(
        worksheet.iter_rows(
            min_row=1,
            max_row=EXCEL_HEADER_SCAN_ROWS,
            max_col=max_column,
            values_only=True,
        ),
        start=1,
    ):
        row_values = list(row)
        score = score_excel_header_row(row_values)

        if score > best_score:
            best_score = score
            best_row_number = row_number
            best_row = row_values

    workbook.close()

    return {
        "rowNumber": best_row_number,
        "rowIndex": best_row_number - 1,
        "columns": make_unique_column_names(best_row),
    }


def clean_dataframe_columns(df: pl.DataFrame) -> pl.DataFrame:
    columns = make_unique_column_names([str(column) for column in df.columns])

    return df.rename(dict(zip(df.columns, columns)))


# Document readers normalize CSV and Excel inputs into Polars dataframes and column lists.
def read_excel_dataframe(
    contents: bytes,
    sheet_name: Optional[str] = None,
    filename: Optional[str] = None,
) -> pl.DataFrame:
    filename_lower = (filename or "").lower()
    read_options = {}

    if not filename_lower or filename_lower.endswith((".xlsx", ".xlsm")):
        try:
            header = detect_excel_header_row(contents, sheet_name)
            read_options["header_row"] = header["rowIndex"]
        except Exception:
            read_options["header_row"] = 0

    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        df = pl.read_excel(
            io.BytesIO(contents),
            sheet_name=sheet_name,
            engine="calamine",
            read_options=read_options or None,
            infer_schema_length=0,
        )

    return clean_dataframe_columns(df)


def read_excel_header_columns(contents: bytes, sheet_name: Optional[str] = None) -> List[str]:
    return detect_excel_header_row(contents, sheet_name)["columns"]


def read_document_columns(contents: bytes, filename: str) -> Dict[str, Any]:
    filename_lower = filename.lower()

    if filename_lower.endswith(".csv"):
        df = pl.read_csv(io.BytesIO(contents), n_rows=1, infer_schema=False)

        return {
            "fileName": filename,
            "fileType": "csv",
            "columns": [str(column) for column in df.columns],
            "sheetNames": [],
            "activeSheet": None,
        }

    if filename_lower.endswith(".xlsx") or filename_lower.endswith(".xls"):
        excel_file = pd.ExcelFile(io.BytesIO(contents))
        sheet_names = excel_file.sheet_names
        first_sheet_name = sheet_names[0]

        try:
            columns = read_excel_header_columns(contents, first_sheet_name)
        except Exception:
            with warnings.catch_warnings():
                warnings.simplefilter("ignore")
                df = pl.read_excel(
                    io.BytesIO(contents),
                    sheet_name=first_sheet_name,
                    engine="calamine",
                    infer_schema_length=0,
                )
            columns = [str(column) for column in df.columns]

        return {
            "fileName": filename,
            "fileType": "excel",
            "columns": columns,
            "sheetNames": sheet_names,
            "activeSheet": first_sheet_name,
        }

    raise HTTPException(
        status_code=400,
        detail="Only CSV and Excel files are supported.",
    )
