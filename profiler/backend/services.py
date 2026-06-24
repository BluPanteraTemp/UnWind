import base64
import io
import json
import math
import re
import struct
import warnings
import zipfile
import zlib
import xml.etree.ElementTree as ET
from typing import Any, Dict, List, Optional
from uuid import uuid4

import pandas as pd
import polars as pl
from fastapi import HTTPException


FILE_STORE = {}

TYPE_DETECTION_SAMPLE_SIZE = 1000
TOP_VALUES_LIMIT = 10
PREVIEW_ROWS = 50
PREVIEW_COLUMN_LIMIT = 100
ISSUE_ROWS_LIMIT = 100
ISSUE_EXAMPLES_LIMIT = 5
M_PREVIEW_LIMIT = 4000
WIDE_PROFILE_COLUMN_LIMIT = 200
EXCEL_HEADER_SCAN_ROWS = 20
EXCEL_HEADER_SCAN_COLS = 1000


warnings.filterwarnings("ignore", message=r".*Could not determine dtype.*")


def xml_local_name(tag: str) -> str:
    return tag.rsplit("}", 1)[-1]


def parse_xml_bytes(contents: bytes) -> ET.Element:
    return ET.fromstring(contents)


def get_xml_attr(element: ET.Element, name: str) -> Optional[str]:
    for key, value in element.attrib.items():
        if xml_local_name(key) == name:
            return value

    return None


def normalize_m_identifier(identifier: str) -> str:
    value = identifier.strip()

    if value.startswith("#\"") and value.endswith("\""):
        return value[2:-1].replace('""', '"')

    return value


def decode_text_part(contents: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-16", "utf-16le", "utf-8"):
        try:
            return contents.decode(encoding)
        except UnicodeDecodeError:
            continue

    return contents.decode("utf-8", errors="ignore")


def read_zip_text(workbook: zipfile.ZipFile, path: str) -> str:
    return decode_text_part(workbook.read(path))


def read_mashup_local_files(payload: bytes) -> Dict[str, bytes]:
    # Excel's DataMashup payload embeds a small ZIP-like package with a short prefix.
    # Python's ZipFile does not always read it cleanly, so walk local file headers directly.
    files = {}
    offset = 0

    while True:
        header_index = payload.find(b"PK\x03\x04", offset)
        if header_index == -1 or header_index + 30 > len(payload):
            break

        try:
            (
                _version,
                _flags,
                compression,
                _modified_time,
                _modified_date,
                _crc,
                compressed_size,
                _uncompressed_size,
                name_length,
                extra_length,
            ) = struct.unpack("<HHHHHIIIHH", payload[header_index + 4:header_index + 30])
        except Exception:
            break

        name_start = header_index + 30
        name_end = name_start + name_length
        data_start = name_end + extra_length
        data_end = data_start + compressed_size

        if data_end > len(payload):
            break

        name = payload[name_start:name_end].decode("utf-8", errors="ignore")
        compressed = payload[data_start:data_end]

        try:
            if compression == 8:
                files[name] = zlib.decompress(compressed, -15)
            elif compression == 0:
                files[name] = compressed
        except zlib.error:
            pass

        offset = data_end

    return files


def extract_mashup_formulas(workbook: zipfile.ZipFile) -> List[Dict[str, str]]:
    formulas = []

    for path in workbook.namelist():
        if not path.startswith("customXml/") or not path.endswith(".xml"):
            continue

        try:
            text = read_zip_text(workbook, path)
        except Exception:
            continue

        if "DataMashup" not in text:
            continue

        try:
            root = ET.fromstring(text)
            encoded_payload = "".join(root.itertext()).strip()
            payload = base64.b64decode(encoded_payload)
        except Exception:
            continue

        for formula_path, contents in read_mashup_local_files(payload).items():
            if formula_path.lower().endswith(".m"):
                formulas.append({
                    "path": formula_path,
                    "code": decode_text_part(contents),
                })

    return formulas


def parse_workbook_connections(workbook: zipfile.ZipFile) -> Dict[str, Dict[str, Any]]:
    if "xl/connections.xml" not in workbook.namelist():
        return {}

    connections = {}
    root = parse_xml_bytes(workbook.read("xl/connections.xml"))

    for connection in root.iter():
        if xml_local_name(connection.tag) != "connection":
            continue

        connection_id = get_xml_attr(connection, "id")
        if not connection_id:
            continue

        name = get_xml_attr(connection, "name") or f"Connection {connection_id}"
        query_name = name.removeprefix("Query - ").strip()
        db_pr = next((child for child in connection if xml_local_name(child.tag) == "dbPr"), None)
        command = get_xml_attr(db_pr, "command") if db_pr is not None else None
        connection_string = get_xml_attr(db_pr, "connection") if db_pr is not None else None

        location_match = re.search(r"Location=(?:&quot;|\")?([^;&\"]+)(?:&quot;|\")?", connection_string or "")
        if location_match:
            query_name = location_match.group(1).strip()

        connections[connection_id] = {
            "id": connection_id,
            "name": name,
            "queryName": query_name,
            "description": get_xml_attr(connection, "description") or "",
            "command": command or "",
            "connectionString": connection_string or "",
            "saveData": get_xml_attr(connection, "saveData") == "1",
            "backgroundRefresh": get_xml_attr(connection, "background") == "1",
        }

    return connections


def parse_workbook_sheets(workbook: zipfile.ZipFile) -> Dict[str, str]:
    if "xl/workbook.xml" not in workbook.namelist() or "xl/_rels/workbook.xml.rels" not in workbook.namelist():
        return {}

    rels_root = parse_xml_bytes(workbook.read("xl/_rels/workbook.xml.rels"))
    rel_targets = {
        get_xml_attr(rel, "Id"): get_xml_attr(rel, "Target")
        for rel in rels_root
        if get_xml_attr(rel, "Id") and get_xml_attr(rel, "Target")
    }
    sheets = {}
    workbook_root = parse_xml_bytes(workbook.read("xl/workbook.xml"))

    for sheet in workbook_root.iter():
        if xml_local_name(sheet.tag) != "sheet":
            continue

        sheet_name = get_xml_attr(sheet, "name")
        rel_id = get_xml_attr(sheet, "id")
        target = rel_targets.get(rel_id)

        if not sheet_name or not target:
            continue

        path = target.lstrip("/")
        if not path.startswith("xl/"):
            path = f"xl/{path}"

        sheets[path] = sheet_name

    return sheets


def normalize_part_path(base_path: str, target: str) -> str:
    if target.startswith("/"):
        return target.lstrip("/")

    parts = base_path.split("/")[:-1] + target.split("/")
    normalized = []

    for part in parts:
        if part in ("", "."):
            continue
        if part == "..":
            if normalized:
                normalized.pop()
            continue
        normalized.append(part)

    return "/".join(normalized)


def parse_relationship_targets(workbook: zipfile.ZipFile, rel_path: str, relation_name: str) -> List[str]:
    if rel_path not in workbook.namelist():
        return []

    root = parse_xml_bytes(workbook.read(rel_path))
    base_path = rel_path.replace("/_rels/", "/").removesuffix(".rels")
    targets = []

    for rel in root:
        rel_type = get_xml_attr(rel, "Type") or ""
        target = get_xml_attr(rel, "Target")

        if relation_name in rel_type and target:
            targets.append(normalize_part_path(base_path, target))

    return targets


def parse_query_outputs(workbook: zipfile.ZipFile, connections: Dict[str, Dict[str, Any]]) -> Dict[str, List[Dict[str, Any]]]:
    sheets_by_path = parse_workbook_sheets(workbook)
    outputs_by_query = {}

    for sheet_path, sheet_name in sheets_by_path.items():
        sheet_number_match = re.search(r"sheet(\d+)\.xml$", sheet_path)
        if not sheet_number_match:
            continue

        rel_path = f"xl/worksheets/_rels/sheet{sheet_number_match.group(1)}.xml.rels"

        for table_path in parse_relationship_targets(workbook, rel_path, "relationships/table"):
            if table_path not in workbook.namelist():
                continue

            try:
                table_root = parse_xml_bytes(workbook.read(table_path))
            except Exception:
                continue

            table_name = get_xml_attr(table_root, "displayName") or get_xml_attr(table_root, "name") or table_path
            table_ref = get_xml_attr(table_root, "ref") or ""

            table_number_match = re.search(r"table(\d+)\.xml$", table_path)
            query_table_paths = []

            if table_number_match:
                query_rel_path = f"xl/tables/_rels/table{table_number_match.group(1)}.xml.rels"
                query_table_paths = parse_relationship_targets(workbook, query_rel_path, "relationships/queryTable")

            for query_table_path in query_table_paths:
                if query_table_path not in workbook.namelist():
                    continue

                try:
                    query_table_root = parse_xml_bytes(workbook.read(query_table_path))
                except Exception:
                    continue

                connection_id = get_xml_attr(query_table_root, "connectionId")
                connection = connections.get(connection_id or "")
                query_name = connection["queryName"] if connection else table_name

                outputs_by_query.setdefault(query_name, []).append({
                    "sheetName": sheet_name,
                    "tableName": table_name,
                    "range": table_ref,
                    "connectionId": connection_id,
                })

    return outputs_by_query


def split_m_queries(formula_code: str) -> List[Dict[str, str]]:
    pattern = re.compile(
        r"(?ms)^shared\s+(?P<name>#\"(?:[^\"]|\"\")*\"|[A-Za-z_][\w.]*)\s*=\s*(?P<body>.*?);(?=\s*(?:shared\s+|$))"
    )

    return [
        {
            "name": normalize_m_identifier(match.group("name")),
            "rawName": match.group("name"),
            "code": match.group("body").strip(),
        }
        for match in pattern.finditer(formula_code)
    ]


def extract_transformations(m_code: str) -> List[Dict[str, str]]:
    known_functions = [
        ("Table.SelectColumns", "Kept selected columns"),
        ("Table.RemoveColumns", "Removed columns"),
        ("Table.RenameColumns", "Renamed columns"),
        ("Table.TransformColumnTypes", "Changed column types"),
        ("Table.SelectRows", "Filtered rows"),
        ("Table.NestedJoin", "Merged queries/tables"),
        ("Table.Join", "Joined queries/tables"),
        ("Table.Combine", "Appended tables"),
        ("Table.Group", "Grouped rows"),
        ("Table.Pivot", "Pivoted values"),
        ("Table.Unpivot", "Unpivoted columns"),
        ("Table.ExpandTableColumn", "Expanded nested table"),
        ("Table.AddColumn", "Added custom column"),
        ("Table.Sort", "Sorted rows"),
        ("Table.Distinct", "Removed duplicates"),
        ("Table.ReplaceValue", "Replaced values"),
    ]

    transformations = []

    for function_name, label in known_functions:
        count = m_code.count(function_name)
        if count:
            transformations.append({
                "function": function_name,
                "label": label,
                "count": count,
            })

    return transformations


def extract_sources(m_code: str) -> List[Dict[str, str]]:
    source_patterns = [
        (r"Excel\.CurrentWorkbook\(\)\{\[Name=\"([^\"]+)\"\]\}", "workbook-table"),
        (r"File\.Contents\(\"([^\"]+)\"\)", "file"),
        (r"Folder\.Files\(\"([^\"]+)\"\)", "folder"),
        (r"Web\.Contents\(\"([^\"]+)\"\)", "web"),
        (r"SharePoint\.Files\(\"([^\"]+)\"", "sharepoint"),
        (r"Sql\.Database\(\"([^\"]+)\"\s*,\s*\"([^\"]+)\"", "sql"),
        (r"Odbc\.DataSource\(\"([^\"]+)\"", "odbc"),
        (r"OData\.Feed\(\"([^\"]+)\"", "odata"),
    ]
    sources = []
    seen = set()

    for pattern, source_type in source_patterns:
        for match in re.finditer(pattern, m_code):
            value = " / ".join(group for group in match.groups() if group)
            key = (source_type, value)

            if key in seen:
                continue

            seen.add(key)
            sources.append({
                "type": source_type,
                "value": value,
            })

    return sources


def extract_query_dependencies(m_code: str, query_names: List[str], current_name: str) -> List[str]:
    dependencies = []

    for query_name in query_names:
        if query_name == current_name:
            continue

        patterns = [
            rf"(?<![\w.]){re.escape(query_name)}(?![\w.])",
            re.escape(f'#{chr(34)}{query_name}{chr(34)}'),
        ]

        if any(re.search(pattern, m_code) for pattern in patterns):
            dependencies.append(query_name)

    return sorted(set(dependencies))


def build_lineage_risks(queries: List[Dict[str, Any]], connections: Dict[str, Dict[str, Any]]) -> List[Dict[str, str]]:
    risks = []
    query_names = {query["name"] for query in queries}

    for connection in connections.values():
        if connection["queryName"] not in query_names:
            risks.append({
                "severity": "warning",
                "message": f"Connection '{connection['name']}' points to query '{connection['queryName']}', but no M definition was extracted.",
            })

    for query in queries:
        if not query["outputs"]:
            risks.append({
                "severity": "info",
                "message": f"Query '{query['name']}' appears to be connection-only or not loaded to a worksheet table.",
            })

        for source in query["sources"]:
            value = source["value"]
            if re.search(r"^[A-Za-z]:\\|/Users/|/home/", value):
                risks.append({
                    "severity": "warning",
                    "message": f"Query '{query['name']}' uses a local path: {value}",
                })

    for query in queries:
        for dependency in query["dependencies"]:
            if dependency not in query_names:
                risks.append({
                    "severity": "warning",
                    "message": f"Query '{query['name']}' references missing query '{dependency}'.",
                })

    return risks


def inspect_workbook_lineage(contents: bytes, filename: str) -> Dict[str, Any]:
    if not filename.lower().endswith((".xlsx", ".xlsm")):
        return {
            "supported": False,
            "queries": [],
            "connections": [],
            "edges": [],
            "risks": [],
            "summary": {
                "queryCount": 0,
                "connectionCount": 0,
                "loadedQueryCount": 0,
                "sourceCount": 0,
                "riskCount": 0,
            },
        }

    try:
        with zipfile.ZipFile(io.BytesIO(contents)) as workbook:
            connections = parse_workbook_connections(workbook)
            outputs_by_query = parse_query_outputs(workbook, connections)
            formula_parts = extract_mashup_formulas(workbook)
    except zipfile.BadZipFile:
        return {
            "supported": False,
            "queries": [],
            "connections": [],
            "edges": [],
            "risks": [{"severity": "warning", "message": "Workbook package could not be inspected."}],
            "summary": {
                "queryCount": 0,
                "connectionCount": 0,
                "loadedQueryCount": 0,
                "sourceCount": 0,
                "riskCount": 1,
            },
        }

    parsed_queries = []

    for formula in formula_parts:
        for query in split_m_queries(formula["code"]):
            query["formulaPath"] = formula["path"]
            parsed_queries.append(query)

    query_names = [query["name"] for query in parsed_queries]
    queries = []
    edges = []

    for query in parsed_queries:
        dependencies = extract_query_dependencies(query["code"], query_names, query["name"])
        sources = extract_sources(query["code"])
        outputs = outputs_by_query.get(query["name"], [])

        for dependency in dependencies:
            edges.append({
                "from": dependency,
                "to": query["name"],
                "type": "query-reference",
            })

        for source in sources:
            edges.append({
                "from": source["value"],
                "to": query["name"],
                "type": source["type"],
            })

        for output in outputs:
            edges.append({
                "from": query["name"],
                "to": f"{output['sheetName']}!{output['tableName']}",
                "type": "loads-to",
            })

        queries.append({
            "name": query["name"],
            "formulaPath": query["formulaPath"],
            "dependencies": dependencies,
            "sources": sources,
            "outputs": outputs,
            "transformations": extract_transformations(query["code"]),
            "mCodePreview": query["code"][:M_PREVIEW_LIMIT],
            "mCodeLength": len(query["code"]),
        })

    risks = build_lineage_risks(queries, connections)
    source_count = sum(len(query["sources"]) for query in queries)

    return {
        "supported": True,
        "queries": queries,
        "connections": list(connections.values()),
        "edges": edges,
        "risks": risks,
        "summary": {
            "queryCount": len(queries),
            "connectionCount": len(connections),
            "loadedQueryCount": sum(1 for query in queries if query["outputs"]),
            "sourceCount": source_count,
            "riskCount": len(risks),
        },
    }


def store_file(contents: bytes, file_type: str, file_name: str) -> str:
    # Keep the raw upload in memory so later requests can re-read it for filters and matrices.
    file_id = str(uuid4())
    FILE_STORE[file_id] = {
        "contents": contents,
        "fileType": file_type,
        "fileName": file_name,
    }
    return file_id


def get_stored_file(file_id: str) -> Dict[str, Any]:
    stored = FILE_STORE.get(file_id)

    if stored is None:
        raise HTTPException(400, "File expired")

    # Older Excel-only entries were stored as raw bytes; keep this fallback during development.
    if isinstance(stored, bytes):
        return {
            "contents": stored,
            "fileType": "excel",
            "fileName": "uploaded-file",
        }

    return stored


def read_stored_dataframe(file_id: str, sheet_name: Optional[str] = None) -> pl.DataFrame:
    # CSVs are read as text first so mixed-type columns do not crash parsing.
    # Numeric/date profiling later uses safe casts where needed.
    stored = get_stored_file(file_id)
    contents = stored["contents"]

    if stored["fileType"] == "csv":
        return pl.read_csv(io.BytesIO(contents), infer_schema=False)

    return read_excel_dataframe(contents, sheet_name, stored["fileName"])


def normalize_excel_header_value(value: Any) -> str:
    if value is None:
        return ""

    if isinstance(value, float) and value.is_integer():
        return str(int(value))

    return str(value).strip()


def make_unique_column_names(values: List[Any]) -> List[str]:
    seen = {}
    columns = []

    for index, value in enumerate(values, start=1):
        name = normalize_excel_header_value(value)

        if not name or re.match(r"^__UNNAMED__\d+$", name):
            name = f"Column {index}"

        count = seen.get(name, 0)
        seen[name] = count + 1

        if count:
            name = f"{name}_{count + 1}"

        columns.append(name)

    return columns


def is_numeric_header_value(value: str) -> bool:
    if not value:
        return False

    return re.fullmatch(r"[-+]?\d+(?:\.\d+)?(?:_\d+)?", value) is not None


def score_excel_header_row(values: List[Any]) -> float:
    normalized = [normalize_excel_header_value(value) for value in values]
    non_empty = [value for value in normalized if value]

    if not non_empty:
        return -1000

    text_like_count = sum(1 for value in non_empty if not is_numeric_header_value(value))
    numeric_count = len(non_empty) - text_like_count
    duplicate_count = len(non_empty) - len(set(non_empty))
    blank_count = len(normalized) - len(non_empty)

    return (
        len(non_empty) * 4
        + text_like_count * 3
        - numeric_count * 2
        - duplicate_count * 4
        - blank_count * 0.15
    )


def detect_excel_header_row(contents: bytes, sheet_name: Optional[str] = None) -> Dict[str, Any]:
    from openpyxl import load_workbook

    workbook = load_workbook(io.BytesIO(contents), read_only=True, data_only=True)
    worksheet = workbook[sheet_name] if sheet_name else workbook[workbook.sheetnames[0]]
    max_column = min(worksheet.max_column or 1, EXCEL_HEADER_SCAN_COLS)
    best_row_number = 1
    best_row = []
    best_score = -1000

    for row_number, row in enumerate(
        worksheet.iter_rows(
            min_row=1,
            max_row=EXCEL_HEADER_SCAN_ROWS,
            max_col=max_column,
            values_only=True,
        ),
        start=1,
    ):
        row_values = list(row)
        score = score_excel_header_row(row_values)

        if score > best_score:
            best_score = score
            best_row_number = row_number
            best_row = row_values

    workbook.close()

    return {
        "rowNumber": best_row_number,
        "rowIndex": best_row_number - 1,
        "columns": make_unique_column_names(best_row),
    }


def clean_dataframe_columns(df: pl.DataFrame) -> pl.DataFrame:
    columns = make_unique_column_names([str(column) for column in df.columns])

    return df.rename(dict(zip(df.columns, columns)))


def read_excel_dataframe(
    contents: bytes,
    sheet_name: Optional[str] = None,
    filename: Optional[str] = None,
) -> pl.DataFrame:
    filename_lower = (filename or "").lower()
    read_options = {}

    if not filename_lower or filename_lower.endswith((".xlsx", ".xlsm")):
        try:
            header = detect_excel_header_row(contents, sheet_name)
            read_options["header_row"] = header["rowIndex"]
        except Exception:
            read_options["header_row"] = 0

    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        df = pl.read_excel(
            io.BytesIO(contents),
            sheet_name=sheet_name,
            engine="calamine",
            read_options=read_options or None,
            infer_schema_length=0,
        )

    return clean_dataframe_columns(df)


def read_excel_header_columns(contents: bytes, sheet_name: Optional[str] = None) -> List[str]:
    return detect_excel_header_row(contents, sheet_name)["columns"]


def read_document_columns(contents: bytes, filename: str) -> Dict[str, Any]:
    filename_lower = filename.lower()

    if filename_lower.endswith(".csv"):
        df = pl.read_csv(io.BytesIO(contents), n_rows=1, infer_schema=False)

        return {
            "fileName": filename,
            "fileType": "csv",
            "columns": [str(column) for column in df.columns],
            "sheetNames": [],
            "activeSheet": None,
        }

    if filename_lower.endswith(".xlsx") or filename_lower.endswith(".xls"):
        excel_file = pd.ExcelFile(io.BytesIO(contents))
        sheet_names = excel_file.sheet_names
        first_sheet_name = sheet_names[0]

        try:
            columns = read_excel_header_columns(contents, first_sheet_name)
        except Exception:
            with warnings.catch_warnings():
                warnings.simplefilter("ignore")
                df = pl.read_excel(
                    io.BytesIO(contents),
                    sheet_name=first_sheet_name,
                    engine="calamine",
                    infer_schema_length=0,
                )
            columns = [str(column) for column in df.columns]

        return {
            "fileName": filename,
            "fileType": "excel",
            "columns": columns,
            "sheetNames": sheet_names,
            "activeSheet": first_sheet_name,
        }

    raise HTTPException(
        status_code=400,
        detail="Only CSV and Excel files are supported.",
    )


def parse_custom_blank_values(raw: Optional[str]) -> List[str]:
    if not raw:
        return []

    return [x.strip().lower() for x in raw.split(",") if x.strip()]


def parse_json_list(raw: Optional[str]) -> List[str]:
    # Frontend sends list-style form values, such as mandatory fields, as JSON strings.
    if not raw:
        return []

    try:
        values = json.loads(raw)
    except json.JSONDecodeError:
        return []

    if not isinstance(values, list):
        return []

    return [str(value) for value in values if str(value).strip()]


def clean_value(value):
    if value is None:
        return None

    if isinstance(value, float):
        if math.isnan(value) or math.isinf(value):
            return None
        return round(value, 2)

    return value


def clean_record(record: Dict[str, Any]) -> Dict[str, Any]:
    return {key: clean_value(value) for key, value in record.items()}


def build_row_preview(row: Dict[str, Any], focus_column: str) -> Dict[str, Any]:
    preview = {}

    if focus_column in row:
        preview[focus_column] = row.get(focus_column)

    for key, value in row.items():
        if key == "_rowNumber" or key == focus_column:
            continue

        preview[key] = value

        if len(preview) >= 6:
            break

    return clean_record(preview)


def is_name_datetime_like(column_name: str) -> bool:
    name = column_name.lower()

    return (
        "date" in name
        or name.endswith("_at")
        or "time" in name
        or "created" in name
        or "updated" in name
        or "valid_from" in name
        or "valid_to" in name
    )


def try_parse_datetime(series: pl.Series) -> pl.Series:
    text = series.cast(pl.Utf8, strict=False)

    return (
        text.str.strptime(pl.Datetime, "%Y-%m-%d %H:%M:%S", strict=False)
        .fill_null(text.str.strptime(pl.Datetime, "%Y-%m-%d", strict=False))
        .fill_null(text.str.strptime(pl.Datetime, "%d/%m/%Y", strict=False))
        .fill_null(text.str.strptime(pl.Datetime, "%m/%d/%Y", strict=False))
    )


def detect_profile_type(column_name: str, series: pl.Series) -> str:
    name = column_name.lower()

    if "email" in name or "mail" in name:
        return "email"

    if "phone" in name or "mobile" in name or "contact" in name:
        return "phone"

    dtype = series.dtype

    numeric_types = {
        pl.Int8, pl.Int16, pl.Int32, pl.Int64,
        pl.UInt8, pl.UInt16, pl.UInt32, pl.UInt64,
        pl.Float32, pl.Float64,
    }

    if dtype in numeric_types:
        return "numeric"

    if dtype == pl.Boolean:
        return "boolean"

    text_series = series.cast(pl.Utf8, strict=False).str.strip_chars()

    non_blank = text_series.filter(
        text_series.is_not_null() & (text_series != "")
    ).head(TYPE_DETECTION_SAMPLE_SIZE)

    values = non_blank.to_list()

    if not values:
        return "empty"

    if is_name_datetime_like(column_name):
        sample = values[:100]

        parsed = try_parse_datetime(pl.Series(sample))

        valid_ratio = parsed.drop_nulls().len() / len(sample)

        if valid_ratio > 0.7:
            return "datetime"

    # CSV columns are loaded as text, so infer numeric-ness from values instead of dtype.
    numeric_probe = pl.Series(values).cast(pl.Float64, strict=False)
    numeric_ratio = numeric_probe.drop_nulls().len() / len(values)

    if numeric_ratio > 0.8:
        return "numeric"

    unique_count = len(set(values))
    unique_ratio = unique_count / len(values)
    avg_length = sum(len(str(v)) for v in values) / len(values)

    if unique_ratio <= 0.3 and avg_length <= 50:
        return "categorical"

    return "text"


def validate_email_text(value: str) -> bool:
    if value is None or str(value).strip() == "":
        return True

    return re.match(r"^[^@\s]+@[^@\s]+\.[^@\s]+$", str(value).strip()) is not None


def validate_phone_text(value: str) -> bool:
    if value is None or str(value).strip() == "":
        return True

    cleaned = re.sub(r"[\s\-\(\)]", "", str(value).strip())

    return (
        cleaned.startswith("+")
        and cleaned[1:].isdigit()
        and 8 <= len(cleaned[1:]) <= 15
    ) or (
        cleaned.isdigit()
        and 8 <= len(cleaned) <= 15
    )


def blank_mask_expr(column_name: str):
    return (
        pl.col(column_name).is_null()
        | (pl.col(column_name).cast(pl.Utf8, strict=False).str.strip_chars() == "")
    )


def custom_blank_mask_expr(column_name: str, custom_blank_values: List[str]):
    if not custom_blank_values:
        return pl.lit(False)

    return (
        pl.col(column_name)
        .cast(pl.Utf8, strict=False)
        .str.strip_chars()
        .str.to_lowercase()
        .is_in(custom_blank_values)
        & ~blank_mask_expr(column_name)
    )


def build_top_values(series: pl.Series):
    s = series.cast(pl.Utf8, strict=False).str.strip_chars()
    s = s.filter(s.is_not_null() & (s != ""))

    if s.is_empty():
        return []

    vc = s.value_counts(sort=True).head(TOP_VALUES_LIMIT)

    name_col = vc.columns[0]
    count_col = vc.columns[1]

    return [
        {"value": str(row[name_col]), "count": int(row[count_col])}
        for row in vc.to_dicts()
    ]


def build_text_stats(series: pl.Series):
    s = series.cast(pl.Utf8, strict=False).str.strip_chars()
    non_empty = s.filter(s.is_not_null() & (s != ""))

    empty_count = int((s == "").sum() or 0)

    if non_empty.is_empty():
        return {
            "minLength": 0,
            "maxLength": 0,
            "meanLength": 0,
            "medianLength": 0,
            "emptyStringCount": empty_count,
        }

    lengths = non_empty.str.len_chars()

    return {
        "minLength": int(lengths.min() or 0),
        "maxLength": int(lengths.max() or 0),
        "meanLength": clean_value(float(lengths.mean() or 0)),
        "medianLength": clean_value(float(lengths.median() or 0)),
        "emptyStringCount": empty_count,
        "avgWordCount": clean_value(
            float(
                non_empty
                .str.split(" ")
                .list.len()
                .mean()
                or 0
            )
        ),
    }


def build_numeric_stats(series: pl.Series):
    numeric = series.cast(pl.Float64, strict=False).drop_nulls()

    if numeric.is_empty():
        return None, set()

    q1 = numeric.quantile(0.25)
    q3 = numeric.quantile(0.75)
    iqr = q3 - q1

    lower = q1 - (1.5 * iqr)
    upper = q3 + (1.5 * iqr)

    outlier_indices = set()

    values = numeric.to_list()
    outlier_count = sum(1 for value in values if value < lower or value > upper)

    hist = numeric.hist(bin_count=10)

    histogram = []
    if hist.height > 0:
        cols = hist.columns
        for row in hist.to_dicts():
            histogram.append(
                {
                    "range": str(row.get(cols[0])),
                    "count": int(row.get(cols[-1]) or 0),
                }
            )

    return {
        "min": clean_value(float(numeric.min())),
        "max": clean_value(float(numeric.max())),
        "mean": clean_value(float(numeric.mean())),
        "median": clean_value(float(numeric.median())),
        "stdDev": clean_value(float(numeric.std() or 0)),
        "q1": clean_value(float(q1)),
        "q3": clean_value(float(q3)),
        "iqr": clean_value(float(iqr)),
        "range": clean_value(float(numeric.max() - numeric.min())),
        "zeroCount": int((numeric == 0).sum() or 0),
        "negativeCount": int((numeric < 0).sum() or 0),
        "outlierCount": int(outlier_count),
        "lowerOutlierBound": clean_value(float(lower)),
        "upperOutlierBound": clean_value(float(upper)),
        "histogram": histogram,
    }, outlier_indices


def build_datetime_stats(series: pl.Series):
    text = series.cast(pl.Utf8, strict=False).str.strip_chars()
    non_blank = text.filter(text.is_not_null() & (text != ""))

    if non_blank.is_empty():
        return None

    parsed = try_parse_datetime(non_blank)

    valid = parsed.drop_nulls()

    if valid.is_empty():
        return {
            "minDate": None,
            "maxDate": None,
            "invalidDateCount": int(non_blank.len()),
        }

    return {
        "minDate": str(valid.min()),
        "maxDate": str(valid.max()),
        "invalidDateCount": int(parsed.is_null().sum() or 0),
    }


def get_recommendation(score_percentage, review_null_above, discard_null_at_least):
    if score_percentage >= discard_null_at_least:
        return "discard"

    if score_percentage > review_null_above:
        return "review"

    return "keep"


def build_column_issue_examples(
    df: pl.DataFrame,
    column_name: str,
    profile_type: str,
    custom_blank_values: List[str],
    statistics: Optional[Dict[str, Any]],
) -> List[Dict[str, Any]]:
    indexed_df = df.with_row_index("_rowNumber", offset=1)
    examples = []
    seen = set()

    def add_rows(issue_type: str, rows: List[Dict[str, Any]]):
        for row in rows:
            key = (int(row.get("_rowNumber", 0)), issue_type)

            if key in seen:
                continue

            seen.add(key)
            examples.append(
                {
                    "rowNumber": int(row.get("_rowNumber", 0)),
                    "issueType": issue_type,
                    "value": clean_value(row.get(column_name)),
                    "rowPreview": build_row_preview(row, column_name),
                }
            )

            if len(examples) >= ISSUE_EXAMPLES_LIMIT:
                return

    blank_rows = (
        indexed_df
        .filter(blank_mask_expr(column_name))
        .head(ISSUE_EXAMPLES_LIMIT)
        .to_dicts()
    )
    add_rows("Blank value", blank_rows)

    if len(examples) < ISSUE_EXAMPLES_LIMIT and custom_blank_values:
        custom_blank_rows = (
            indexed_df
            .filter(custom_blank_mask_expr(column_name, custom_blank_values))
            .head(ISSUE_EXAMPLES_LIMIT)
            .to_dicts()
        )
        add_rows("Custom blank value", custom_blank_rows)

    if len(examples) < ISSUE_EXAMPLES_LIMIT and profile_type in ["email", "phone"]:
        rows = (
            indexed_df
            .select(["_rowNumber", *df.columns])
            .to_dicts()
        )
        invalid_rows = []

        for row in rows:
            value = row.get(column_name)
            text_value = "" if value is None else str(value).strip()

            if text_value == "":
                continue

            if profile_type == "email" and not validate_email_text(text_value):
                invalid_rows.append(row)

            if profile_type == "phone" and not validate_phone_text(text_value):
                invalid_rows.append(row)

            if len(invalid_rows) >= ISSUE_EXAMPLES_LIMIT:
                break

        label = "Invalid email format" if profile_type == "email" else "Invalid phone format"
        add_rows(label, invalid_rows)

    if len(examples) < ISSUE_EXAMPLES_LIMIT and profile_type == "datetime":
        text_expr = pl.col(column_name).cast(pl.Utf8, strict=False).str.strip_chars()
        parsed_expr = (
            text_expr.str.strptime(pl.Datetime, "%Y-%m-%d %H:%M:%S", strict=False)
            .fill_null(text_expr.str.strptime(pl.Datetime, "%Y-%m-%d", strict=False))
            .fill_null(text_expr.str.strptime(pl.Datetime, "%d/%m/%Y", strict=False))
            .fill_null(text_expr.str.strptime(pl.Datetime, "%m/%d/%Y", strict=False))
        )
        invalid_date_rows = (
            indexed_df
            .filter(text_expr.is_not_null() & (text_expr != "") & parsed_expr.is_null())
            .head(ISSUE_EXAMPLES_LIMIT)
            .to_dicts()
        )
        add_rows("Invalid date value", invalid_date_rows)

    if len(examples) < ISSUE_EXAMPLES_LIMIT and profile_type == "numeric" and statistics:
        lower = statistics.get("lowerOutlierBound")
        upper = statistics.get("upperOutlierBound")

        if lower is not None and upper is not None:
            numeric_expr = pl.col(column_name).cast(pl.Float64, strict=False)
            outlier_rows = (
                indexed_df
                .filter(numeric_expr.is_not_null() & ((numeric_expr < lower) | (numeric_expr > upper)))
                .head(ISSUE_EXAMPLES_LIMIT)
                .to_dicts()
            )
            add_rows("Possible outlier", outlier_rows)

    return examples


def build_column_valid_example(
    df: pl.DataFrame,
    column_name: str,
    profile_type: str,
    custom_blank_values: List[str],
    statistics: Optional[Dict[str, Any]],
) -> Optional[Dict[str, Any]]:
    rows = df.with_row_index("_rowNumber", offset=1).to_dicts()
    lower = statistics.get("lowerOutlierBound") if statistics else None
    upper = statistics.get("upperOutlierBound") if statistics else None

    for row in rows:
        value = row.get(column_name)
        text_value = "" if value is None else str(value).strip()

        if text_value == "" or text_value.lower() in custom_blank_values:
            continue

        if profile_type == "email" and not validate_email_text(text_value):
            continue

        if profile_type == "phone" and not validate_phone_text(text_value):
            continue

        if profile_type == "datetime":
            parsed = try_parse_datetime(pl.Series([text_value]))

            if parsed.drop_nulls().is_empty():
                continue

        if profile_type == "numeric":
            numeric_value = pl.Series([value]).cast(pl.Float64, strict=False).item()

            if numeric_value is None:
                continue

            if lower is not None and upper is not None and (numeric_value < lower or numeric_value > upper):
                continue

        return {
            "rowNumber": int(row.get("_rowNumber", 0)),
            "value": clean_value(value),
        }

    return None


def profile_column(
    df: pl.DataFrame,
    column_name: str,
    row_count: int,
    custom_blank_values: List[str],
    review_null_above: float,
    discard_null_at_least: float,
    include_custom_blanks: bool,
    is_mandatory: bool,
    include_examples: bool = True,
):
    series = df[column_name]

    profile_type = detect_profile_type(column_name, series)

    counts = df.select(
        [
            blank_mask_expr(column_name).sum().alias("true_blank_count"),
            custom_blank_mask_expr(column_name, custom_blank_values).sum().alias("custom_blank_count"),
        ]
    ).to_dicts()[0]

    true_blank_count = int(counts["true_blank_count"] or 0)
    custom_blank_count = int(counts["custom_blank_count"] or 0)

    true_blank_percentage = round((true_blank_count / row_count) * 100, 2) if row_count else 0
    custom_blank_percentage = round((custom_blank_count / row_count) * 100, 2) if row_count else 0

    recommendation_score_percentage = true_blank_percentage

    if include_custom_blanks:
        recommendation_score_percentage = min(100, true_blank_percentage + custom_blank_percentage)

    usable = (
        df
        .filter(~blank_mask_expr(column_name) & ~custom_blank_mask_expr(column_name, custom_blank_values))
        .select(pl.col(column_name).cast(pl.Utf8, strict=False).str.strip_chars())
        .to_series()
    )

    unique_count = int(usable.n_unique()) if not usable.is_empty() else 0
    duplicate_count = int(row_count - unique_count)

    issues = []

    if true_blank_percentage > 25:
        issues.append("High missing values")

    if custom_blank_count > 0:
        issues.append("Contains custom blank values")

    if is_mandatory and true_blank_count + custom_blank_count > 0:
        issues.append("Mandatory field has blank values")

    if unique_count == 1 and row_count > 1:
        issues.append("Only one unique value")

    invalid_count = 0
    invalid_percentage = 0

    if profile_type in ["email", "phone"]:
        text_values = series.cast(pl.Utf8, strict=False).str.strip_chars()
        values = text_values.filter(
            text_values.is_not_null() & (text_values != "")
        ).to_list()

        if profile_type == "email":
            invalid_count = sum(1 for value in values if not validate_email_text(value))
            if invalid_count:
                issues.append("Invalid email format")

        if profile_type == "phone":
            invalid_count = sum(1 for value in values if not validate_phone_text(value))
            if invalid_count:
                issues.append("Invalid phone format")

        invalid_percentage = round((invalid_count / len(values)) * 100, 2) if values else 0

    statistics = None
    outlier_indices = set()

    if profile_type == "numeric":
        statistics, outlier_indices = build_numeric_stats(series)

        if statistics:
            if statistics["outlierCount"] > 0:
                issues.append("Contains possible outliers")

            if statistics["negativeCount"] > 0:
                issues.append("Contains negative values")

    elif profile_type == "datetime":
        statistics = build_datetime_stats(series)

        if statistics and statistics["invalidDateCount"] > 0:
            issues.append("Invalid date values")

    else:
        statistics = build_text_stats(series)

    recommendation = get_recommendation(
        recommendation_score_percentage,
        review_null_above,
        discard_null_at_least,
    )
    issue_examples = []
    valid_example = None

    if include_examples:
        issue_examples = build_column_issue_examples(
            df=df,
            column_name=column_name,
            profile_type=profile_type,
            custom_blank_values=custom_blank_values,
            statistics=statistics,
        )
        valid_example = build_column_valid_example(
            df=df,
            column_name=column_name,
            profile_type=profile_type,
            custom_blank_values=custom_blank_values,
            statistics=statistics,
        )

    return {
        "name": column_name,
        "isMandatory": is_mandatory,
        "profileType": profile_type,
        "quality": {
            "nullCount": true_blank_count,
            "nullPercentage": true_blank_percentage,
            "customBlankCount": custom_blank_count,
            "customBlankPercentage": custom_blank_percentage,
            "recommendationScorePercentage": round(recommendation_score_percentage, 2),
            "uniqueCount": unique_count,
            "duplicateCount": duplicate_count,
            "invalidCount": invalid_count,
            "invalidPercentage": invalid_percentage,
        },
        "statistics": statistics,
        "topValues": build_top_values(series),
        "issues": issues,
        "issueExamples": issue_examples,
        "validExample": valid_example,
        "recommendation": recommendation,
    }


def limit_row_record_columns(record: Dict[str, Any], output_columns: List[str]) -> Dict[str, Any]:
    limited = {}

    if "_rowNumber" in record:
        limited["_rowNumber"] = record["_rowNumber"]

    for column in output_columns:
        if column in record:
            limited[column] = record[column]

    return limited


def build_issue_rows(
    df: pl.DataFrame,
    columns_profile: List[Dict[str, Any]],
    custom_blank_values: List[str],
    output_columns: Optional[List[str]] = None,
):
    # Return a small preview of rows with blank/custom blank issues for quick inspection.
    issue_expr = pl.lit(False)

    for col in columns_profile:
        name = col["name"]

        issue_expr = issue_expr | blank_mask_expr(name) | custom_blank_mask_expr(name, custom_blank_values)

    issue_df = df.with_row_index("_rowNumber", offset=1).filter(issue_expr).head(ISSUE_ROWS_LIMIT)

    records = []

    for row in issue_df.to_dicts():
        reasons = []

        for col in columns_profile:
            name = col["name"]
            label = f"{name} *" if col.get("isMandatory") else name
            value = row.get(name)

            if value is None or str(value).strip() == "":
                reasons.append(f"{label}: blank")
            elif str(value).strip().lower() in custom_blank_values:
                reasons.append(f"{label}: custom blank")

        output_row = limit_row_record_columns(row, output_columns or list(row.keys()))
        output_row["_issueReasons"] = reasons
        records.append(clean_record(output_row))

    return records


def profile_dataframe(
    df: pl.DataFrame,
    file_name: str,
    custom_blank_values: List[str],
    mandatory_fields: List[str],
    review_null_above: float,
    discard_null_at_least: float,
    include_custom_blanks: bool,
):
    # This is the central profiling pipeline used by CSVs, Excel sheets, and re-profile actions.
    df = df.rename({name: str(name) for name in df.columns})
    mandatory_field_set = set(mandatory_fields)

    row_count = df.height
    column_count = df.width
    wide_profile_mode = column_count > WIDE_PROFILE_COLUMN_LIMIT
    preview_columns = df.columns[:PREVIEW_COLUMN_LIMIT] if wide_profile_mode else df.columns

    columns = [
        profile_column(
            df=df,
            column_name=column_name,
            row_count=row_count,
            custom_blank_values=custom_blank_values,
            review_null_above=review_null_above,
            discard_null_at_least=discard_null_at_least,
            include_custom_blanks=include_custom_blanks,
            is_mandatory=column_name in mandatory_field_set,
            include_examples=not wide_profile_mode,
        )
        for column_name in df.columns
    ]

    duplicate_count = int(df.is_duplicated().sum() or 0)

    duplicate_rows = [
        clean_record(row)
        for row in (
            df
            .with_row_index("_rowNumber", offset=1)
            .filter(df.is_duplicated())
            .head(ISSUE_ROWS_LIMIT)
            .to_dicts()
        )
    ]
    duplicate_rows = [limit_row_record_columns(row, preview_columns) for row in duplicate_rows]

    for row in duplicate_rows:
        row["_issueReasons"] = ["Duplicate row"]

    issue_rows = build_issue_rows(df, columns, custom_blank_values, preview_columns)

    recommendation_counts = {
        "keep": sum(1 for col in columns if col["recommendation"] == "keep"),
        "review": sum(1 for col in columns if col["recommendation"] == "review"),
        "discard": sum(1 for col in columns if col["recommendation"] == "discard"),
    }

    total_true_blank_cells = sum(col["quality"]["nullCount"] for col in columns)
    total_custom_blank_cells = sum(col["quality"]["customBlankCount"] for col in columns)

    preview = [
        clean_record(row)
        for row in df.select(preview_columns).head(PREVIEW_ROWS).to_dicts()
    ]

    return {
        "dataset": {
            "fileName": file_name,
            "rowCount": row_count,
            "columnCount": column_count,
            "duplicateRows": duplicate_count,
        },
        "settings": {
            "reviewNullAbove": review_null_above,
            "discardNullAtLeast": discard_null_at_least,
            "includeCustomBlanks": include_custom_blanks,
            "customBlankValues": custom_blank_values,
            "mandatoryFields": mandatory_fields,
        },
        "summary": {
            "recommendationCounts": recommendation_counts,
            "columnsWithIssues": sum(1 for col in columns if len(col["issues"]) > 0),
            "mandatoryColumns": sum(1 for col in columns if col["isMandatory"]),
            "wideProfileMode": wide_profile_mode,
            "wideProfileColumnLimit": WIDE_PROFILE_COLUMN_LIMIT,
            "totalTrueBlankCells": int(total_true_blank_cells),
            "totalCustomBlankCells": int(total_custom_blank_cells),
            "totalBlankCells": int(total_true_blank_cells + total_custom_blank_cells),
            "issueRowsReturned": len(issue_rows),
            "duplicateRowsReturned": len(duplicate_rows),
        },
        "columns": columns,
        "preview": preview,
        "issueRows": issue_rows,
        "duplicateRowsPreview": duplicate_rows,
    }


def apply_row_filters(df: pl.DataFrame, raw_filters: Optional[str]) -> pl.DataFrame:
    # Row filters keep only records where selected columns match selected values.
    if not raw_filters:
        return df

    filters = json.loads(raw_filters)

    if not filters:
        return df

    expr = pl.lit(True)

    for f in filters:
        column = f.get("column")
        values = f.get("values", [])

        if not column or column not in df.columns or not values:
            continue

        normalized_values = [str(v).strip() for v in values]

        expr = expr & (
            pl.col(column)
            .cast(pl.Utf8, strict=False)
            .str.strip_chars()
            .is_in(normalized_values)
        )

    return df.filter(expr)


def get_matrix_recommendation(blank_percentage, review_null_above, discard_null_at_least):
    if blank_percentage >= discard_null_at_least:
        return "discard"

    if blank_percentage > review_null_above:
        return "review"

    return "keep"


def build_matrix(
    df: pl.DataFrame,
    group_by: str,
    custom_blank_values: List[str],
    review_null_above: float,
    discard_null_at_least: float,
    include_custom_blanks: bool,
    max_groups: int = 30,
):
    if group_by not in df.columns:
        raise HTTPException(400, "Invalid group by column")

    group_values_df = (
        df
        .select(
            pl.col(group_by)
            .cast(pl.Utf8, strict=False)
            .str.strip_chars()
            .alias(group_by)
        )
        .filter(pl.col(group_by).is_not_null() & (pl.col(group_by) != ""))
        .group_by(group_by)
        .len()
        .sort("len", descending=True)
        .head(max_groups)
    )

    group_values = [row[group_by] for row in group_values_df.to_dicts()]

    fields = [col for col in df.columns if col != group_by]

    rows = []

    for field in fields:
        cells = []

        for group_value in group_values:
            group_df = df.filter(
                pl.col(group_by)
                .cast(pl.Utf8, strict=False)
                .str.strip_chars()
                == group_value
            )

            total = group_df.height

            if total == 0:
                blank_count = 0
                custom_blank_count = 0
                blank_percentage = 0
            else:
                counts = group_df.select(
                    [
                        blank_mask_expr(field).sum().alias("blank_count"),
                        custom_blank_mask_expr(field, custom_blank_values)
                        .sum()
                        .alias("custom_blank_count"),
                    ]
                ).to_dicts()[0]

                blank_count = int(counts["blank_count"] or 0)
                custom_blank_count = int(counts["custom_blank_count"] or 0)

                score_count = blank_count

                if include_custom_blanks:
                    score_count += custom_blank_count

                blank_percentage = round((score_count / total) * 100, 2)

            recommendation = get_matrix_recommendation(
                blank_percentage,
                review_null_above,
                discard_null_at_least,
            )

            cells.append(
                {
                    "groupValue": group_value,
                    "totalRows": total,
                    "blankCount": blank_count,
                    "customBlankCount": custom_blank_count,
                    "blankPercentage": blank_percentage,
                    "recommendation": recommendation,
                }
            )

        rows.append(
            {
                "field": field,
                "cells": cells,
            }
        )

    return {
        "groupBy": group_by,
        "groups": [
            {
                "value": row[group_by],
                "rowCount": row["len"],
            }
            for row in group_values_df.to_dicts()
        ],
        "rows": rows,
    }
